import re
from openpyxl import Workbook
from openpyxl import load_workbook

#Imp: make it crash free
while True:
    try:
        wb_name=input('workbook name?')
        wb=load_workbook(wb_name)
        break
    except:
        print('You have not put a valid excel file name, or this code does not exist in the same folder.')
        continue

#Imp: make it crash free
while True:
    try:
        sh=input('sheet name?')
        ws=wb[sh]
        break
    except:
        print('No such sheet exists!')
        continue

# target_column1=input('Target column 1?')
# target_column2=input('Target column 2?')

while True:
    try:
        column=input('which column or column range?')
        col=ws[column]
        break
    except:
        print('Not a valid column name')
        continue
# tcol1=ws[target_column1]
# tcol2=ws[target_column2]

for cell in col:
    try:
        data=str(cell.value)
        val=re.findall('[a-zA-Z]',data)
        if data.strip()=='None':
            continue
        # val=val.strip()
        without_space=''
        for letters in val:
            without_space=without_space+letters.strip()
        cell.value=without_space
    except:
        continue

while True:
    try:
        wb.save(wb_name)
        print('done')
        break
    except:
        while True:
            print('Please close the file first.')
            closed=input('Closed[y/n]?')
            if closed.lower()=='y':
                break
            else:
                continue
