import re
from openpyxl import Workbook
from openpyxl import load_workbook


while True:
    try:
        wb_name=input('workbook name?')
        wb=load_workbook(wb_name)
        break
    except:
        print('You have not put a valid excel file name, or this code does not exist in the same folder.')
        continue


while True:
    try:
        sh=input('sheet name?')
        ws=wb[sh]
        break
    except:
        print('No such sheet exists!')
        continue


while True:
    try:
        column=input('which column or column range?')
        col=ws[column]
        break
    except:
        print('Not a valid column name')
        continue

while True:
    try:
        target_column1=input('Target column 1?')
        tcol1=ws[target_column1]
        break
    except:
        print('Not a valid column name')
        continue

while True:
    try:
        target_column2=input('Target column 2?')
        tcol2=ws[target_column2]
        break
    except:
        print('Not a valid column name')
        continue



for cell,tcell1,tcell2 in zip(col,tcol1,tcol2):
    val=str(cell.value).strip()
    # print(val,type(val))
    if val=='None':
        continue
    # val=val.strip()
    try:
        val_text= re.findall('[a-zA-Z]+',val)
        string=''
        for letters in val_text:
            string=string+' '+letters
        string=string.strip()
        # print(string)

        tcell1.value=string
    except:
        pass

    try:
        val_other=re.findall('[^a-zA-Z]+',val)
        characters=''
        for char in val_other:
            characters=characters+' '+char
        characters=characters.strip()
        # print(characters)

        tcell2.value=characters
    except:
        continue

while True:
    try:
        wb.save(wb_name)
        print('done')
        break
    except:
        while True:
            print('Please close the file first.')
            closed=input('Closed[y/n]?')
            if closed.lower()=='y':
                break
            else:
                continue
